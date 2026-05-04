"""budget_reader.py — Read and parse Juan's financial dashboard from local Excel files.

Reads Juan_Financial_Dashboard.xlsx from BUDGET_FOLDER (configurable via .env).
Returns structured dicts for downstream analysis. Never makes assumptions about
missing data — callers must handle None values explicitly.
"""

import os
import csv
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------- #
# Config
# --------------------------------------------------------------------------- #

_DEFAULT_FOLDER = "/mnt/c/Users/jbsan/OneDrive/Documents/Payoff debts/Payoff debts"
BUDGET_FOLDER = Path(os.getenv("BUDGET_FOLDER", _DEFAULT_FOLDER))

_DASHBOARD_NAME = "Juan_Financial_Dashboard.xlsx"


# --------------------------------------------------------------------------- #
# Internal helpers
# --------------------------------------------------------------------------- #

def _load_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        return None


def _safe_float(val) -> Optional[float]:
    """Convert a cell value to float; return None on failure."""
    if val is None:
        return None
    try:
        f = float(val)
        return f if not (f != f) else None  # NaN guard
    except (TypeError, ValueError):
        return None


def _pct(val) -> Optional[float]:
    """Excel stores APR as 0.2874 meaning 28.74%. Return as percentage float."""
    f = _safe_float(val)
    if f is None:
        return None
    # Values < 1 are stored as decimals; convert to whole-number percent
    return round(f * 100, 4) if f < 1 else round(f, 4)


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #

def get_budget_folder() -> Path:
    return BUDGET_FOLDER


def find_dashboard() -> Optional[Path]:
    """Return path to the dashboard file, or None if not found."""
    p = BUDGET_FOLDER / _DASHBOARD_NAME
    if p.exists():
        return p
    # Also scan for any .xlsx or .csv in the folder
    if BUDGET_FOLDER.exists():
        for f in BUDGET_FOLDER.iterdir():
            if f.suffix.lower() in (".xlsx", ".xls") and f.name.startswith("Juan"):
                return f
    return None


def read_debts() -> dict:
    """Parse the Debts sheet and return structured data.

    Returns:
        {
            "ok": bool,
            "error": str | None,
            "debts": [
                {
                    "priority": int,
                    "name": str,
                    "balance": float | None,
                    "apr_pct": float | None,   # e.g. 28.74 (not 0.2874)
                    "min_payment": float | None,
                    "monthly_interest": float | None,
                    "status": str | None,
                    "notes": str | None,
                }
            ],
            "total_debt": float | None,
            "total_min_payments": float | None,
            "monthly_interest_total": float | None,
        }
    """
    openpyxl = _load_openpyxl()
    if openpyxl is None:
        return {"ok": False, "error": "openpyxl not installed — run: pip3 install openpyxl --break-system-packages", "debts": []}

    dashboard = find_dashboard()
    if dashboard is None:
        return {"ok": False, "error": f"Budget file not found in {BUDGET_FOLDER}", "debts": []}

    try:
        wb = openpyxl.load_workbook(str(dashboard), data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Could not open budget file: {e}", "debts": []}

    # --- Debts sheet ---
    if "Debts" not in wb.sheetnames:
        return {"ok": False, "error": "Debts sheet not found in workbook", "debts": []}

    ws = wb["Debts"]
    rows = list(ws.iter_rows(values_only=True))

    debts = []
    total_debt = None
    total_min = None
    monthly_interest_total = None

    for row in rows:
        # Skip empty rows and header/section rows
        if not row or row[0] is None:
            continue

        # Data rows have a numeric priority in column 0
        if isinstance(row[0], (int, float)) and not isinstance(row[0], bool):
            priority = int(row[0])
            name = str(row[1]).strip() if row[1] else f"Debt #{priority}"
            debts.append({
                "priority": priority,
                "name": name,
                "min_payment": _safe_float(row[2]),
                "balance": _safe_float(row[3]),
                "apr_pct": _pct(row[4]),
                "monthly_interest": _safe_float(row[5]),
                "status": str(row[7]).strip() if row[7] else None,
                "notes": str(row[8]).strip() if row[8] else None,
            })

        # KEY METRICS section rows
        if row[0] == "Total Debt (excl. mortgage)":
            total_debt = _safe_float(row[1])
        elif row[0] == "Total Monthly Min Payments":
            total_min = _safe_float(row[1])
        elif row[0] == "Monthly Interest Burned":
            monthly_interest_total = _safe_float(row[1])

    return {
        "ok": True,
        "error": None,
        "debts": debts,
        "total_debt": total_debt,
        "total_min_payments": total_min,
        "monthly_interest_total": monthly_interest_total,
    }


def read_snapshot() -> dict:
    """Parse the Snapshot sheet for high-level financial summary.

    Returns:
        {
            "ok": bool,
            "error": str | None,
            "monthly_income": float | None,
            "total_expenses": float | None,
            "net_cash_flow": float | None,
            "total_debt": float | None,
            "monthly_interest": float | None,
            "debt_to_income": float | None,  # as decimal e.g. 0.75
        }
    """
    openpyxl = _load_openpyxl()
    if openpyxl is None:
        return {"ok": False, "error": "openpyxl not installed", "monthly_income": None}

    dashboard = find_dashboard()
    if dashboard is None:
        return {"ok": False, "error": f"Budget file not found in {BUDGET_FOLDER}", "monthly_income": None}

    try:
        wb = openpyxl.load_workbook(str(dashboard), data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Could not open budget file: {e}", "monthly_income": None}

    # Sheet name contains emoji
    snap_sheet = next((s for s in wb.sheetnames if "Snapshot" in s), None)
    if snap_sheet is None:
        return {"ok": False, "error": "Snapshot sheet not found", "monthly_income": None}

    ws = wb[snap_sheet]
    rows = list(ws.iter_rows(values_only=True))

    # Row 3 (index 2) has headers; row 4 (index 3) has values
    # Monthly Income, Total Expenses, Net Cash Flow, (skip), Total Debt, Monthly Interest, Debt-to-Income
    income = expenses = net = total_debt = monthly_int = dti = None
    for i, row in enumerate(rows):
        if row and row[0] == "Monthly Income":
            # Next row is values
            if i + 1 < len(rows):
                vrow = rows[i + 1]
                income = _safe_float(vrow[0])
                expenses = _safe_float(vrow[1])
                net = _safe_float(vrow[2])
                total_debt = _safe_float(vrow[4])
                monthly_int = _safe_float(vrow[5])
                dti = _safe_float(vrow[6])
            break

    return {
        "ok": True,
        "error": None,
        "monthly_income": income,
        "total_expenses": expenses,
        "net_cash_flow": net,
        "total_debt": total_debt,
        "monthly_interest": monthly_int,
        "debt_to_income": dti,
    }


def read_calendar() -> dict:
    """Parse the Calendar sheet for monthly payment schedule.

    Returns:
        {
            "ok": bool,
            "error": str | None,
            "payments": [
                {"due_date": str, "name": str, "amount": float | None, "category": str | None, "autopay": str | None}
            ]
        }
    """
    openpyxl = _load_openpyxl()
    if openpyxl is None:
        return {"ok": False, "error": "openpyxl not installed", "payments": []}

    dashboard = find_dashboard()
    if dashboard is None:
        return {"ok": False, "error": f"Budget file not found in {BUDGET_FOLDER}", "payments": []}

    try:
        wb = openpyxl.load_workbook(str(dashboard), data_only=True)
    except Exception as e:
        return {"ok": False, "error": f"Could not open budget file: {e}", "payments": []}

    cal_sheet = next((s for s in wb.sheetnames if "Calendar" in s), None)
    if cal_sheet is None:
        return {"ok": False, "error": "Calendar sheet not found", "payments": []}

    ws = wb[cal_sheet]
    rows = list(ws.iter_rows(values_only=True))

    payments = []
    header_seen = False
    for row in rows:
        if not row or row[0] is None:
            continue
        if row[0] == "Due Date":
            header_seen = True
            continue
        if not header_seen:
            continue
        # Row: due_date, name, amount, category, autopay, account, notes
        due = str(row[0]).strip() if row[0] else None
        name = str(row[1]).strip() if row[1] else None
        if not due or not name:
            continue
        payments.append({
            "due_date": due,
            "name": name,
            "amount": _safe_float(row[2]),
            "category": str(row[3]).strip() if row[3] else None,
            "autopay": str(row[4]).strip() if row[4] else None,
        })

    return {"ok": True, "error": None, "payments": payments}
